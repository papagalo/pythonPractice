from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
wb = Workbook()

# load existing spreadsheet
wb = load_workbook('PythonPractice.xlsx')

# Create an active worksheet
ws = wb.active

# Set a variable
#name = ws["A2"].value

# Print something from our spreadsheet
#print(f'{name}:{ws["B2"].value}')

# Grab a whole column
#column_a = ws['A']

# For loop
#for cell in column_a:
#	print(cell.value)

# Grab a range
range = ws['A2':'B5']
print(range)

for cell in range:
	for x in cell:
		print(x.value)

# Changing something